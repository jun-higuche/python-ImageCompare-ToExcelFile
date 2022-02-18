import cv2, os, pathlib, math, shutil
import openpyxl as px
from openpyxl.styles import PatternFill
from opencv_japanese import imread, imwrite
from opencv_compareImg import createDiffImg, imageCompareNum

PIX_PER_CELL_HEIGH = 18 #エクセルの1セルの高さのピクセル数
PIX_PER_CELL_WIDTH = 72 #エクセルの1セルの幅のピクセル数

#画像の縮小割合
IMAGE_RESIZE_PER = 0.4

#このソースコードが存在するディレクトリ
dirname =  os.path.dirname(__file__)

targetDir1 = dirname + "/tmp01"
targetDir2 = dirname + "/tmp02"
saveDir = dirname + "/saveDir"

#縮小画像を一時的に格納するディレクトリ名（自動的に生成して後で削除）
tmpDir1 = dirname + "/shukuTmp01"
tmpDir2 = dirname + "/shukuTmp02"
tmpSaveDir = dirname + "/shukuTmpSaveDir"

#対象の画像ファイル一覧（１）
path_list1 = list(pathlib.Path(targetDir1).glob(r"*.png"))
#対象の画像ファイル一覧（２）
path_list2 = list(pathlib.Path(targetDir2).glob(r"*.png"))

# ２画像の差分画像を出力するディレクトリを生成
os.makedirs(saveDir, exist_ok=True)

#保存するExcelファイル名（新規で作成される）
savefile = dirname + "/save.xlsx"

#類似度
ruiziDict = {}

#ここから画像比較　類似度の算出と差分画像の生成
print("画像比較処理")
for i, path in enumerate(path_list1):
    
    print(str(path))
    img = px.drawing.image.Image(path)

    ruizido = 0.0
    for t, path2 in enumerate(path_list2):

        if path.name == path2.name:
            img2 = px.drawing.image.Image(path2)

            # ２画像の類似度を算出して一時保存
            ruizido = imageCompareNum(str(path), str(path2))
            ruiziDict.update([(path.name, ruizido)])
            
            # ２画像の差分画像を生成
            saveName = saveDir + "\\D_" + path.name
            createDiffImg(str(path), str(path2), saveName)
            break


#ここから画像の縮小処理
print("画像縮小処理")
os.makedirs(tmpDir1, exist_ok=True)
os.makedirs(tmpDir2, exist_ok=True)
os.makedirs(tmpSaveDir, exist_ok=True)

path_list1 = list(pathlib.Path(targetDir1).glob(r"*.png"))
path_list2 = list(pathlib.Path(targetDir2).glob(r"*.png"))
path_list3 = list(pathlib.Path(saveDir).glob(r"*.png"))

for i, path in enumerate(path_list1):
    print(str(path))
    image = imread(str(path))
    h, w = image.shape[:2]
    resized_image = cv2.resize(image, (int(w * IMAGE_RESIZE_PER), int(h * IMAGE_RESIZE_PER)))
    imwrite(tmpDir1 + "/" + path.name, resized_image)

for i, path in enumerate(path_list2):
    print(str(path))
    image = imread(str(path))
    h, w = image.shape[:2]
    resized_image = cv2.resize(image, (int(w * IMAGE_RESIZE_PER), int(h * IMAGE_RESIZE_PER)))
    imwrite(tmpDir2 + "/" + path.name, resized_image)

for i, path in enumerate(path_list3):
    print(str(path))
    image = imread(str(path))
    h, w = image.shape[:2]
    resized_image = cv2.resize(image, (int(w * IMAGE_RESIZE_PER), int(h * IMAGE_RESIZE_PER)))
    imwrite(tmpSaveDir + "/" + path.name, resized_image)


#ここから画像をExcelに貼り付ける処理
print("Excelへの画像貼り付け処理")

wb = px.Workbook()
wb.create_sheet(index=0, title='比較結果０１')
ws = wb['比較結果０１']
wb.remove(wb['Sheet'])

path_list1 = list(pathlib.Path(tmpDir1).glob(r"*.png"))
path_list2 = list(pathlib.Path(tmpDir2).glob(r"*.png"))
path_list3 = list(pathlib.Path(tmpSaveDir).glob(r"*.png"))

paste_cell = 3
tmpWidth = 0

for i, path in enumerate(path_list1):
    
    # 画像読み出し
    print(str(path))
    img = px.drawing.image.Image(path)

    # ファイル名を出力
    ws[px.utils.get_column_letter(1) + str(paste_cell-1)] = path.name

    # 画像を貼り付ける位置をExcel形式で表す(※列のアルファベット+行の番号)
    pos = px.utils.get_column_letter(1) #エクセルの1列目（A列）を指定
    pos = pos + str(paste_cell) #各ファイルの画像を重ならないように貼り付ける

    # 画像を追加
    ws.add_image(img, pos)

    # 次の画像の横軸の場所を画像の幅から計算
    tmpImage2Pos_retu = px.utils.get_column_letter(math.ceil((img.width / PIX_PER_CELL_WIDTH)) *2 + 1)
    tmpImage2Pos = tmpImage2Pos_retu + str(paste_cell)
    tmpCompareImgPos_retu = px.utils.get_column_letter(math.ceil((img.width / PIX_PER_CELL_WIDTH) + 1))
    tmpCompareImgPos = tmpCompareImgPos_retu + str(paste_cell)

    ruizido = 0.0
    for t, path2 in enumerate(path_list2):

        if path.name == path2.name:
            img2 = px.drawing.image.Image(path2)

            # ２画像の類似度を書き込む
            ruizido = ruiziDict[path.name]
            ws[tmpCompareImgPos_retu + str(paste_cell-1)] = "類似度：" + str(ruizido)

            # ２画像の差分画像を貼り付ける
            saveName = tmpSaveDir + "\\D_" + path.name
            img3 = px.drawing.image.Image(saveName)
            ws.add_image(img3, tmpCompareImgPos)
            ws.add_image(img2, tmpImage2Pos)
            break
            
    # 類似度から背景色を生成
    haikeiNoudo = int((float(int('ff', 16)) * ruizido))
    haikeiNoudo = format(haikeiNoudo, 'x')
    haikei = "FF" + str(haikeiNoudo).zfill(2) + str(haikeiNoudo).zfill(2)
    tmpFill = PatternFill(patternType="solid", fgColor=haikei)

    # 類似度の数値に合わせて背景（赤）を設定
    for rows in ws["A" + str(paste_cell-1) : "AK" + str(paste_cell-1)]:
        for cell in rows:
            cell.fill = tmpFill

    # 画像の高さ[pixel]情報から次に貼り付ける画像のセル位置を算出
    paste_cell += math.ceil((img.height / PIX_PER_CELL_HEIGH) + 2)


#Excelデータを保存
wb.save(savefile)

#一時ファイルの削除
shutil.rmtree(tmpDir1)
shutil.rmtree(tmpDir2)
shutil.rmtree(tmpSaveDir)

#処理が完了したことをユーザーに知らせる
print("Excelファイルの作成が完了しました")
